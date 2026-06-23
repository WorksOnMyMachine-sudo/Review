"""Build summary workbook from ChatGPT-classified review template."""

from __future__ import annotations

import argparse
import math
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "data" / "output"

REQUIRED_COLUMNS = {
    "ReviewID",
    "机型ID",
    "机型名称",
    "Channel",
    "来源站点",
    "评分",
    "评论日期",
    "标题",
    "评论内容",
    "问题摘要",
    "一级分类_请填写",
    "二级分类_请填写",
}


def main() -> None:
    args = parse_args()
    input_path = resolve_input(args.input)
    output_path = resolve_output(args.output)

    reviews = load_classified_reviews(input_path)
    try:
        write_summary_workbook(reviews, output_path, input_path)
    except PermissionError:
        fallback_path = unlocked_fallback_path(output_path)
        safe_print(
            f"WARNING: 输出文件可能正在被 Excel 打开，无法覆盖；已改为保存到: {fallback_path}",
            stream=sys.stderr,
        )
        write_summary_workbook(reviews, fallback_path, input_path)
        output_path = fallback_path

    safe_print(f"INPUT: {input_path}")
    safe_print(f"OUTPUT: {output_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Summarize ChatGPT-classified reviews by model and primary issue category."
    )
    parser.add_argument(
        "--input",
        "-i",
        type=Path,
        help="Completed classification template. Defaults to latest review_classification_template*_completed.xlsx.",
    )
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        help="Output summary Excel. Defaults to data/output/review_classification_summary_<timestamp>.xlsx.",
    )
    return parser.parse_args()


def resolve_input(path: Path | None) -> Path:
    if path:
        resolved = path if path.is_absolute() else ROOT / path
        if not resolved.exists():
            raise FileNotFoundError(resolved)
        return resolved

    candidates = sorted(
        OUTPUT_DIR.glob("review_classification_template*_completed.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(
            f"No completed classification template found in {OUTPUT_DIR}"
        )
    return candidates[0]


def resolve_output(path: Path | None) -> Path:
    if path:
        return path if path.is_absolute() else ROOT / path
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return OUTPUT_DIR / f"review_classification_summary_{stamp}.xlsx"


def unlocked_fallback_path(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return path.with_name(f"{path.stem}_{stamp}{path.suffix}")


def load_classified_reviews(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="待分类Reviews")
    missing = REQUIRED_COLUMNS - set(df.columns)
    if missing:
        raise ValueError(f"Missing expected columns: {sorted(missing)}")

    df = df.copy()
    df["一级分类"] = df["一级分类_请填写"].fillna("").astype(str).str.strip()
    df["二级分类"] = df["二级分类_请填写"].fillna("").astype(str).str.strip()
    df["一级分类"] = df["一级分类"].replace("", "Other")
    df["二级分类"] = df["二级分类"].replace("", "Other")
    df["尺寸"] = df.apply(extract_size, axis=1)
    df["评论日期"] = df["评论日期"].map(format_date)
    return df


def extract_size(row: pd.Series) -> str:
    existing_size = row.get("尺寸")
    if pd.notna(existing_size) and str(existing_size).strip():
        return str(existing_size).strip()

    sizes = "32|40|43|50|55|58|65|70|75|85|98|100"
    source_pattern = re.compile(rf"(?:^|_)({sizes})(?:$|_)")
    model_prefix_pattern = re.compile(rf"^({sizes})(?=[A-Z])", re.IGNORECASE)
    text_pattern = re.compile(
        rf"(?<!\d)({sizes})\s*(?:[\"”]|inch|inches|class|吋|寸)",
        re.IGNORECASE,
    )

    for column, pattern in (
        ("来源站点", source_pattern),
        ("机型ID", model_prefix_pattern),
        ("机型名称", model_prefix_pattern),
    ):
        value = row.get(column)
        text = "" if pd.isna(value) else str(value)
        match = pattern.search(text)
        if match:
            return f"{match.group(1)}寸"

    for column in ("标题", "评论内容", "问题摘要"):
        value = row.get(column)
        text = "" if pd.isna(value) else str(value)
        match = text_pattern.search(text)
        if match:
            return f"{match.group(1)}寸"

    return "未知"


def format_date(value) -> str:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return "" if pd.isna(value) else str(value)
    return parsed.strftime("%Y-%m-%d")


def write_summary_workbook(df: pd.DataFrame, output_path: Path, source_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        overview = build_overview(df)
        overview.to_excel(writer, sheet_name="总表", index=False)

        for model_id, model_df in df.groupby("机型ID", sort=False):
            write_model_sheet(writer, str(model_id), model_df, source_path)

        format_workbook(writer.book)


def build_overview(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for (model_id, size), model_df in df.groupby(["机型ID", "尺寸"], sort=False):
        model_name = first_non_empty(model_df["机型名称"])
        total = len(model_df)
        counts = model_df["一级分类"].value_counts()
        for category, count in counts.items():
            rows.append(
                {
                    "机型ID": model_id,
                    "机型名称": model_name,
                    "尺寸": size,
                    "一级分类": category,
                    "问题数": int(count),
                    "占比": count / total if total else 0,
                    "低分评论总数": total,
                }
            )
    return pd.DataFrame(rows)


def write_model_sheet(
    writer: pd.ExcelWriter,
    model_id: str,
    model_df: pd.DataFrame,
    source_path: Path,
) -> None:
    total = len(model_df)
    primary_summary = (
        model_df.groupby(["尺寸", "一级分类"], sort=False)
        .size()
        .reset_index(name="问题数")
        .sort_values(["尺寸", "问题数"], ascending=[True, False], kind="stable")
    )
    primary_summary["占比"] = primary_summary["问题数"] / total if total else 0
    primary_summary["低分评论总数"] = total

    secondary_summary = (
        model_df.groupby(["尺寸", "一级分类", "二级分类"], sort=False)
        .size()
        .reset_index(name="问题数")
        .sort_values(["尺寸", "一级分类", "问题数"], ascending=[True, True, False], kind="stable")
    )
    secondary_summary["占比"] = secondary_summary["问题数"] / total if total else 0

    detail_cols = [
        "机型名称",
        "尺寸",
        "Channel",
        "整体评分",
        "评分",
        "评论日期",
        "一级分类",
        "二级分类",
        "评论内容",
        "分类理由_可选",
    ]
    detail_cols = [col for col in detail_cols if col in model_df.columns]
    details = model_df[detail_cols].copy()
    details["_sort_date"] = pd.to_datetime(details["评论日期"], errors="coerce")
    details = details.sort_values(
        by="_sort_date",
        ascending=False,
        na_position="last",
        kind="stable",
    ).drop(columns=["_sort_date"])
    details.insert(0, "序号", range(1, len(details) + 1))

    sheet_name = safe_sheet_name(model_id)
    header = pd.DataFrame(
        [
            [f"{model_id} 分类汇总"],
            [f"来源文件: {source_path.name}"],
            [f"低分评论数: {total}"],
        ]
    )
    header.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=0)
    primary_summary.to_excel(writer, sheet_name=sheet_name, index=False, startrow=4)
    secondary_start = 7 + len(primary_summary)
    secondary_summary.to_excel(writer, sheet_name=sheet_name, index=False, startrow=secondary_start)
    details_start = secondary_start + len(secondary_summary) + 3
    details.to_excel(writer, sheet_name=sheet_name, index=False, startrow=details_start)


def first_non_empty(series: pd.Series) -> str:
    for value in series:
        if pd.notna(value) and str(value).strip():
            return str(value).strip()
    return ""


def safe_sheet_name(name: str, max_len: int = 31) -> str:
    invalid = set(r"[]:*?/\\")
    cleaned = "".join(c if c not in invalid else "_" for c in name).strip() or "sheet"
    return cleaned[:max_len]


def format_workbook(workbook) -> None:
    hisense_green = "00979B"
    header_fill = PatternFill("solid", fgColor=hisense_green)
    thin_side = Side(style="thin", color="B7C9CC")
    table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    body_font = Font(name="Microsoft YaHei", size=11)
    header_font = Font(name="Microsoft YaHei", size=11, color="FFFFFF", bold=True)
    title_font = Font(name="Microsoft YaHei", size=14, bold=True, color=hisense_green)

    for sheet in workbook.worksheets:
        if sheet.title != "总表" and sheet.max_column > 1:
            sheet.merge_cells(
                start_row=1,
                start_column=1,
                end_row=1,
                end_column=sheet.max_column,
            )
            sheet.cell(row=1, column=1).font = title_font
            sheet.cell(row=1, column=1).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            sheet.freeze_panes = "A5"
        else:
            sheet.freeze_panes = "A2"

        for row in sheet.iter_rows():
            for cell in row:
                cell.font = body_font
                existing_horizontal = cell.alignment.horizontal
                cell.alignment = Alignment(
                    horizontal=existing_horizontal,
                    vertical="center",
                    wrap_text=True,
                )
                if is_header_cell(cell.value):
                    cell.fill = header_fill
                    cell.font = header_font

        center_columns_by_header(sheet, {"一级分类", "二级分类", "问题数", "占比"})
        apply_number_formats(sheet)
        apply_table_borders(sheet, table_border)

        for column in range(1, sheet.max_column + 1):
            letter = get_column_letter(column)
            if column == 1:
                sheet.column_dimensions[letter].width = 14
            else:
                sheet.column_dimensions[letter].width = column_width(sheet, column)

        set_row_heights(sheet)


def is_header_cell(value) -> bool:
    return value in {
        "机型名称",
        "机型ID",
        "尺寸",
        "一级分类",
        "二级分类",
        "问题数",
        "占比",
        "低分评论总数",
        "序号",
        "Channel",
        "整体评分",
        "评分",
        "评论日期",
        "标题",
        "评论内容",
        "问题摘要",
        "分类理由_可选",
    }


def center_columns_by_header(sheet, headers_to_center: set[str]) -> None:
    for row in sheet.iter_rows():
        for header_cell in row:
            if header_cell.value not in headers_to_center:
                continue
            header_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            for column_cells in sheet.iter_cols(
                min_col=header_cell.column,
                max_col=header_cell.column,
                min_row=header_cell.row + 1,
                max_row=sheet.max_row,
            ):
                for cell in column_cells:
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True,
                    )


def apply_number_formats(sheet) -> None:
    for row in sheet.iter_rows():
        headers = {cell.column: cell.value for cell in row if isinstance(cell.value, str)}
        data_start = row[0].row + 1
        data_end = table_data_end_row(sheet, data_start)
        for column, header in headers.items():
            if header == "占比":
                for column_cells in sheet.iter_cols(
                    min_col=column,
                    max_col=column,
                    min_row=data_start,
                    max_row=data_end,
                ):
                    for cell in column_cells:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "0.0%"
            elif header in {"序号", "问题数", "低分评论总数"}:
                for column_cells in sheet.iter_cols(
                    min_col=column,
                    max_col=column,
                    min_row=data_start,
                    max_row=data_end,
                ):
                    for cell in column_cells:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "0"
            elif header == "评分":
                for column_cells in sheet.iter_cols(
                    min_col=column,
                    max_col=column,
                    min_row=data_start,
                    max_row=data_end,
                ):
                    for cell in column_cells:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "0.0"


def table_data_end_row(sheet, data_start: int) -> int:
    row_idx = data_start
    while row_idx <= sheet.max_row:
        values = [sheet.cell(row=row_idx, column=col).value for col in range(1, sheet.max_column + 1)]
        if all(value is None for value in values):
            return row_idx - 1
        row_idx += 1
    return sheet.max_row


def apply_table_borders(sheet, border: Border) -> None:
    """Add borders only to contiguous table regions, leaving spacer rows clean."""
    for row in sheet.iter_rows():
        if not any(is_header_cell(cell.value) for cell in row):
            continue
        populated_columns = [cell.column for cell in row if cell.value is not None]
        if not populated_columns:
            continue
        start_column = min(populated_columns)
        end_column = max(populated_columns)
        end_row = table_data_end_row(sheet, row[0].row + 1)
        for cells in sheet.iter_rows(
            min_row=row[0].row,
            max_row=end_row,
            min_col=start_column,
            max_col=end_column,
        ):
            for cell in cells:
                cell.border = border


def column_width(sheet, column: int) -> int:
    values = [
        sheet.cell(row=row, column=column).value
        for row in range(1, min(sheet.max_row, 80) + 1)
    ]
    max_len = max((len(str(value)) for value in values if value is not None), default=8)
    return max(10, min(max_len + 2, 60))


def set_row_heights(sheet) -> None:
    """Expand rows enough to display long review text without altering cell contents."""
    for row in range(1, sheet.max_row + 1):
        values = [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]
        longest = max((len(str(value)) for value in values if value is not None), default=0)
        explicit_lines = max(
            (str(value).count("\n") + 1 for value in values if value is not None),
            default=1,
        )
        estimated_lines = max(explicit_lines, math.ceil(longest / 58))
        sheet.row_dimensions[row].height = min(409, max(24, estimated_lines * 15))


def safe_print(value: str, *, stream=None) -> None:
    target = stream or sys.stdout
    try:
        print(value, file=target)
    except UnicodeEncodeError:
        encoded = value.encode(target.encoding or "utf-8", errors="backslashreplace")
        print(encoded.decode(target.encoding or "utf-8"), file=target)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:  # noqa: BLE001
        safe_print(f"ERROR: {exc}", stream=sys.stderr)
        raise
