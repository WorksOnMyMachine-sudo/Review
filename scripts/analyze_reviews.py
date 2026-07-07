"""Analyze low-rating reviews by model and export issue summaries to Excel."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
from collections import Counter
from copy import copy
from datetime import datetime
from pathlib import Path

import pandas as pd
import httpx
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "data" / "output"
DEFAULT_CATEGORY_FILE = OUTPUT_DIR / "Call log分类.xlsx"
load_dotenv(ROOT / ".env", override=True)

REQUIRED_COLUMNS = {
    "机型ID",
    "机型名称",
    "Channel",
    "来源站点",
    "来源URL",
    "评分",
    "标题",
    "评论内容",
    "评论日期",
    "用户",
}

FALLBACK_ISSUE_RULES: list[tuple[str, list[str]]] = [
    (
        "画质 / 屏幕显示",
        [
            "picture",
            "image",
            "display",
            "screen",
            "panel",
            "color",
            "brightness",
            "contrast",
            "black level",
            "backlight",
            "blooming",
            "glare",
            "reflection",
            "matte",
            "pixel",
            "dead pixel",
            "hdr",
            "qled",
        ],
    ),
    (
        "软件 / 系统 / App",
        [
            "software",
            "firmware",
            "update",
            "os",
            "google tv",
            "fire tv",
            "app",
            "apps",
            "streaming",
            "netflix",
            "youtube",
            "crash",
            "freeze",
            "lag",
            "slow",
            "bug",
            "glitch",
            "restart",
            "reboot",
        ],
    ),
    (
        "连接 / 端口 / 网络",
        [
            "wifi",
            "wi-fi",
            "internet",
            "network",
            "bluetooth",
            "hdmi",
            "arc",
            "earc",
            "usb",
            "port",
            "connect",
            "connection",
            "disconnect",
            "pairing",
            "signal",
        ],
    ),
    (
        "声音 / 音频",
        [
            "sound",
            "audio",
            "speaker",
            "volume",
            "bass",
            "dialog",
            "voice",
            "mute",
            "surround",
            "soundbar",
        ],
    ),
    (
        "遥控器 / 操作体验",
        [
            "remote",
            "button",
            "menu",
            "interface",
            "navigation",
            "control",
            "input",
            "settings",
            "voice control",
        ],
    ),
    (
        "安装 / 挂装 / 设置",
        [
            "setup",
            "install",
            "installation",
            "mount",
            "wall mount",
            "stand",
            "assemble",
            "assembly",
            "instructions",
            "manual",
            "calibration",
        ],
    ),
    (
        "质量 / 故障 / 损坏",
        [
            "defect",
            "defective",
            "broken",
            "damage",
            "damaged",
            "dead",
            "stopped working",
            "doesn't work",
            "not working",
            "failed",
            "failure",
            "issue",
            "problem",
            "returned",
            "replacement",
        ],
    ),
    (
        "配送 / 包装",
        [
            "delivery",
            "delivered",
            "shipping",
            "ship",
            "box",
            "package",
            "packaging",
            "arrived",
            "carrier",
            "fedex",
            "ups",
        ],
    ),
    (
        "客服 / 保修 / 退换货",
        [
            "customer service",
            "support",
            "warranty",
            "repair",
            "refund",
            "return",
            "exchange",
            "best buy",
            "hisense support",
        ],
    ),
    (
        "价格 / 价值感",
        [
            "price",
            "expensive",
            "cheap",
            "value",
            "worth",
            "money",
            "cost",
            "deal",
        ],
    ),
]

STOPWORDS = {
    "a",
    "an",
    "and",
    "are",
    "as",
    "be",
    "but",
    "by",
    "can",
    "could",
    "for",
    "from",
    "has",
    "have",
    "in",
    "is",
    "it",
    "no",
    "not",
    "of",
    "on",
    "or",
    "other",
    "please",
    "some",
    "the",
    "this",
    "to",
    "tv",
    "with",
}


def main() -> None:
    args = parse_args()
    run_analysis(args)


def run_analysis(args: argparse.Namespace) -> None:
    input_path = resolve_input(args.input)
    output_path = resolve_output(args.output)
    category_path = resolve_category_file(args.category_file)
    issue_rules = load_issue_rules(category_path)

    reviews, dedupe_stats = load_reviews(input_path)
    write_deduped_reviews(reviews, input_path)
    low_reviews = reviews[reviews["评分_num"].le(3)].copy()
    low_reviews["问题摘要"] = low_reviews.apply(make_issue_summary, axis=1)
    if args.classification_mode == "export-template":
        template_path = resolve_template_output(args.template_output)
        write_classification_template(low_reviews, issue_rules, template_path, input_path, category_path)
        safe_print(f"INPUT: {input_path}")
        safe_print(f"DEDUPED_REVIEWS: {input_path}")
        safe_print(f"CATEGORY: {category_path or 'built-in fallback rules'}")
        safe_print("CLASSIFICATION_MODE: export-template")
        safe_print(f"TEMPLATE_OUTPUT: {template_path}")
        return

    if args.classification_mode == "semantic":
        issue_matches = pd.Series(
            classify_reviews_semantic(low_reviews, issue_rules),
            index=low_reviews.index,
        )
    else:
        issue_matches = low_reviews.apply(lambda row: classify_review_keyword(row, issue_rules), axis=1)
    low_reviews["问题分类"] = issue_matches.map(lambda item: item[0])
    low_reviews["二级问题分类"] = issue_matches.map(lambda item: item[1])
    low_reviews["分类理由"] = issue_matches.map(
        lambda item: item[2] if len(item) > 2 else ""
    )
    low_reviews["问题摘要中文"] = translate_summaries(low_reviews["问题摘要"].tolist())

    write_analysis(low_reviews, reviews, output_path, input_path, dedupe_stats, issue_rules, category_path)
    safe_print(f"INPUT: {input_path}")
    safe_print(f"DEDUPED_REVIEWS: {input_path}")
    safe_print(f"CATEGORY: {category_path or 'built-in fallback rules'}")
    safe_print(f"CLASSIFICATION_MODE: {args.classification_mode}")
    safe_print(f"OUTPUT: {output_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Analyze <=3-star reviews and export issue summaries by model."
    )
    parser.add_argument(
        "--input",
        "-i",
        type=Path,
        help="Input review Excel. Defaults to data/output/reviews_incremental.xlsx.",
    )
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        help="Output analysis Excel. Defaults to data/output/review_issue_analysis_latest.xlsx.",
    )
    parser.add_argument(
        "--category-file",
        "-c",
        type=Path,
        help="Call log category Excel. Defaults to data/output/Call log分类.xlsx.",
    )
    parser.add_argument(
        "--classification-mode",
        choices=["semantic", "keyword", "export-template"],
        default="semantic",
        help=(
            "semantic uses OpenAI API to understand full review meaning; "
            "keyword uses local fallback matching; export-template creates a ChatGPT upload template."
        ),
    )
    parser.add_argument(
        "--template-output",
        type=Path,
        help="Output Excel for --classification-mode export-template.",
    )
    return parser.parse_args()


def resolve_input(path: Path | None) -> Path:
    if path:
        path = path if path.is_absolute() else ROOT / path
        if not path.exists():
            raise FileNotFoundError(path)
        return path

    incremental_path = OUTPUT_DIR / "reviews_incremental.xlsx"
    if incremental_path.exists():
        return incremental_path

    candidates = sorted(
        (
            path
            for path in OUTPUT_DIR.glob("reviews_*.xlsx")
            if "__deduped_tmp" not in path.stem and not path.stem.endswith("_deduped")
        ),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(f"No reviews_*.xlsx found in {OUTPUT_DIR}")
    return candidates[0]


def resolve_output(path: Path | None) -> Path:
    if path:
        return path if path.is_absolute() else ROOT / path
    return OUTPUT_DIR / "review_issue_analysis_latest.xlsx"


def resolve_template_output(path: Path | None) -> Path:
    if path:
        return path if path.is_absolute() else ROOT / path
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return OUTPUT_DIR / f"review_classification_template_{stamp}.xlsx"


def resolve_category_file(path: Path | None) -> Path | None:
    if path:
        resolved = path if path.is_absolute() else ROOT / path
        if not resolved.exists():
            raise FileNotFoundError(resolved)
        return resolved
    return DEFAULT_CATEGORY_FILE if DEFAULT_CATEGORY_FILE.exists() else None


def load_issue_rules(category_path: Path | None) -> list[tuple[str, list[str]]]:
    if not category_path:
        return FALLBACK_ISSUE_RULES

    workbook = pd.ExcelFile(category_path)
    rules: list[tuple[str, list[str]]] = []
    for sheet_name in workbook.sheet_names:
        df = pd.read_excel(category_path, sheet_name=sheet_name)
        for column in df.columns:
            category = str(column).strip()
            if not category or category.lower().startswith("unnamed"):
                continue
            phrases = [
                str(value).strip()
                for value in df[column].dropna().tolist()
                if str(value).strip()
            ]
            if phrases:
                rules.append((category, phrases))

    return rules or FALLBACK_ISSUE_RULES


def load_reviews(path: Path) -> tuple[pd.DataFrame, dict[str, int]]:
    workbook = pd.ExcelFile(path)
    frames = []
    for sheet in workbook.sheet_names:
        df = pd.read_excel(path, sheet_name=sheet)
        if df.empty or not REQUIRED_COLUMNS.issubset(df.columns):
            continue
        frames.append(df)

    if not frames:
        raise ValueError(f"No review sheets with expected columns found in {path}")

    data = pd.concat(frames, ignore_index=True)
    original_count = len(data)
    data = dedupe_reviews(data)
    deduped_count = len(data)
    data["评分_num"] = data["评分"].map(parse_rating)
    stats = {
        "original_count": original_count,
        "deduped_count": deduped_count,
        "removed_count": original_count - deduped_count,
    }
    return data[data["评分_num"].notna()].copy(), stats


def dedupe_reviews(df: pd.DataFrame) -> pd.DataFrame:
    dedupe_cols = [
        "机型ID",
        "Channel",
        "来源站点",
        "标题",
        "评论内容",
        "评论日期",
        "用户",
    ]
    out = df.copy()
    key_cols = []
    for col in dedupe_cols:
        key_col = f"_dedupe_{col}"
        key_cols.append(key_col)
        out[key_col] = out[col].map(normalize_dedupe_value)
    out = out.drop_duplicates(subset=key_cols, keep="first").copy()
    return out.drop(columns=key_cols)


def write_deduped_reviews(df: pd.DataFrame, output_path: Path) -> None:
    export_df = df.drop(columns=["评分_num"], errors="ignore").copy()
    temp_path = output_path.with_name(f"{output_path.stem}__deduped_tmp{output_path.suffix}")

    with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
        for model_id, group in export_df.groupby("机型ID", sort=False):
            sheet_name = safe_sheet_name(str(model_id))
            group.to_excel(writer, sheet_name=sheet_name, index=False)
        format_deduped_review_workbook(writer.book)

    try:
        temp_path.replace(output_path)
    except PermissionError:
        fallback_path = output_path.with_name(f"{output_path.stem}_deduped{output_path.suffix}")
        temp_path.replace(fallback_path)
        safe_print(
            f"WARNING: 原始 reviews 文件可能正被 Excel 打开，无法覆盖；已保存去重副本: {fallback_path}",
            stream=sys.stderr,
        )


def write_classification_template(
    df: pd.DataFrame,
    issue_rules: list[tuple[str, list[str]]],
    output_path: Path,
    source_path: Path,
    category_path: Path | None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    template_df = df.copy()
    if "整体评分" not in template_df.columns:
        template_df["整体评分"] = ""
    template_df["尺寸"] = template_df.apply(extract_size, axis=1)
    template_df["_sort_date"] = template_df["评论日期"].map(parse_review_date)
    template_df["评论日期"] = template_df["评论日期"].map(format_review_date)
    template_df = template_df.sort_values(
        by=["机型ID", "尺寸", "_sort_date"],
        ascending=[True, True, False],
        na_position="last",
        kind="stable",
    ).reset_index(drop=True)
    template_df.insert(0, "ReviewID", [f"R{i + 1:04d}" for i in range(len(template_df))])
    template_df["一级分类_请填写"] = ""
    template_df["二级分类_请填写"] = ""
    template_df["分类理由_可选"] = ""

    review_cols = [
        "ReviewID",
        "机型ID",
        "机型名称",
        "尺寸",
        "Channel",
        "来源站点",
        "整体评分",
        "评分",
        "评论日期",
        "标题",
        "评论内容",
        "问题摘要",
        "一级分类_请填写",
        "二级分类_请填写",
        "分类理由_可选",
    ]
    reviews_to_classify = template_df[review_cols].copy()

    taxonomy_rows = []
    for primary, secondaries in issue_rules:
        for secondary in secondaries:
            taxonomy_rows.append({"一级分类": primary, "二级分类": secondary})
    taxonomy_df = pd.DataFrame(taxonomy_rows)

    instructions = pd.DataFrame(
        [
            ["用途", "把本文件上传到 ChatGPT 网页版，让 ChatGPT 阅读完整 review 语义后填写分类。"],
            ["来源Review文件", source_path.name],
            ["Call log分类表", category_path.name if category_path else "内置兜底规则"],
            ["分类要求", "不要只按关键词匹配。请理解用户真实抱怨点，从 Call log 分类表中选择最合适的一级分类和二级分类。"],
            ["多问题处理", "如果一条评论包含多个问题，选择最能解释低评分的主要问题。"],
            ["无有效正文", "如果 review 没有可用抱怨内容，一级分类和二级分类都填 Other。"],
            ["输出要求", "请在“待分类Reviews”sheet 中填写“一级分类_请填写”和“二级分类_请填写”，不要改 ReviewID、机型和尺寸。"],
        ],
        columns=["项目", "说明"],
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        instructions.to_excel(writer, sheet_name="使用说明", index=False)
        reviews_to_classify.to_excel(writer, sheet_name="待分类Reviews", index=False)
        taxonomy_df.to_excel(writer, sheet_name="Call log分类表", index=False)
        format_classification_template_workbook(writer.book)


def format_classification_template_workbook(workbook) -> None:
    header_fill = PatternFill("solid", fgColor="00979B")
    thin_side = Side(style="thin", color="B7C9CC")
    table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    body_font = Font(name="Microsoft YaHei", size=11)
    header_font = Font(name="Microsoft YaHei", size=11, color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            set_cell_alignment(cell, vertical="center", wrap_text=True)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                set_cell_alignment(cell, vertical="center", wrap_text=True)
        for column in range(1, sheet.max_column + 1):
            letter = get_column_letter(column)
            sheet.column_dimensions[letter].width = column_width(sheet, column)
        apply_used_range_borders(sheet, table_border)


def format_deduped_review_workbook(workbook) -> None:
    header_fill = PatternFill("solid", fgColor="00979B")
    thin_side = Side(style="thin", color="B7C9CC")
    table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    body_font = Font(name="Microsoft YaHei", size=11)
    header_font = Font(name="Microsoft YaHei", size=11, color="FFFFFF", bold=True)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            set_cell_alignment(cell, vertical="center", wrap_text=True)

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                set_cell_alignment(cell, vertical="center", wrap_text=True)

        for column in range(1, sheet.max_column + 1):
            letter = get_column_letter(column)
            sheet.column_dimensions[letter].width = column_width(sheet, column)
        apply_used_range_borders(sheet, table_border)


def normalize_dedupe_value(value) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip().lower()


def parse_rating(value) -> float | None:
    if pd.isna(value):
        return None
    match = re.search(r"\d+(?:\.\d+)?", str(value))
    return float(match.group(0)) if match else None


def parse_review_date(value) -> pd.Timestamp | pd.NaT:
    if pd.isna(value):
        return pd.NaT
    text = str(value).replace("\xa0", " ").strip()
    match = re.search(r"on\s+(.+)$", text, re.I)
    if match:
        text = match.group(1).strip()
    parsed = pd.to_datetime(text, errors="coerce")
    return parsed if not pd.isna(parsed) else pd.NaT


def format_review_date(value) -> str:
    parsed = parse_review_date(value)
    if pd.isna(parsed):
        return ""
    return parsed.strftime("%Y-%m-%d")


def extract_size(row: pd.Series) -> str:
    existing_size = row.get("尺寸")
    if pd.notna(existing_size) and str(existing_size).strip():
        return str(existing_size).strip()

    sizes = "32|40|43|50|55|58|65|70|75|85|98|100"
    source_pattern = re.compile(rf"(?:^|_)({sizes})(?:$|_)")
    model_prefix_pattern = re.compile(rf"^({sizes})(?=[A-Z])", re.IGNORECASE)
    text_pattern = re.compile(
        rf"(?<!\d)({sizes})\s*(?:[\"”]|inch|inches|class|吋|寸)",
        re.IGNORECASE,
    )

    for column, pattern in (
        ("来源站点", source_pattern),
        ("机型ID", model_prefix_pattern),
        ("机型名称", model_prefix_pattern),
    ):
        value = row.get(column)
        text = "" if pd.isna(value) else str(value)
        match = pattern.search(text)
        if match:
            return f"{match.group(1)}寸"

    for column in ("标题", "评论内容", "问题摘要"):
        value = row.get(column)
        text = "" if pd.isna(value) else str(value)
        match = text_pattern.search(text)
        if match:
            return f"{match.group(1)}寸"

    return "未知"


def classify_reviews_semantic(
    df: pd.DataFrame,
    issue_rules: list[tuple[str, list[str]]],
) -> list[tuple[str, str, str]]:
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError(
            "语义分类需要 OPENAI_API_KEY。请在 .env 中配置可用 key，"
            "或手动添加 --classification-mode keyword 使用关键词备用模式。"
        )

    model = os.getenv("REVIEW_ANALYSIS_CLASSIFY_MODEL", "gpt-4o-mini").strip()
    batch_size = read_int_env("REVIEW_ANALYSIS_CLASSIFY_BATCH_SIZE", default=3, minimum=1, maximum=30)
    taxonomy = {category: phrases for category, phrases in issue_rules}
    rows = review_rows_for_semantic_classification(df)
    total = len(rows)
    cache = load_semantic_cache()
    results: list[tuple[str, str, str] | None] = [None] * total
    pending_rows: list[dict] = []

    for row in rows:
        cached = cache.get(review_cache_key(row))
        if cached:
            results[row["id"]] = tuple(cached)
        else:
            pending_rows.append(row)

    cached_count = total - len(pending_rows)
    if cached_count:
        safe_print(f"语义分类缓存: 已复用 {cached_count}/{total} 条")

    completed = cached_count
    for batch in chunked(pending_rows, batch_size):
        safe_print(f"语义分类进度: {completed}/{total} 条，正在请求 {len(batch)} 条")
        batch_results = classify_batch_semantic(
            batch,
            taxonomy=taxonomy,
            api_key=api_key,
            model=model,
        )
        for row, item in zip(batch, batch_results):
            normalized = tuple(item)
            results[row["id"]] = normalized
            cache[review_cache_key(row)] = list(normalized)
        save_semantic_cache(cache)
        completed += len(batch)
        safe_print(f"语义分类进度: {completed}/{total} 条")

    return [
        item if item is not None else ("Other", "Please add comments", "未生成分类结果")
        for item in results
    ][: len(df)]


def review_rows_for_semantic_classification(df: pd.DataFrame) -> list[dict]:
    rows = []
    for item_id, (_, row) in enumerate(df.iterrows()):
        rows.append(
            {
                "id": item_id,
                "cache_id": make_review_stable_id(row),
                "model": str(row.get("机型名称", "")),
                "channel": str(row.get("Channel", "")),
                "rating": str(row.get("评分", "")),
                "date": str(row.get("评论日期", "")),
                "title": truncate_text(row.get("标题", ""), 160),
                "review": truncate_text(row.get("评论内容", ""), 900),
                "summary": truncate_text(row.get("问题摘要", ""), 260),
            }
        )
    return rows


def classify_batch_semantic(
    rows: list[dict],
    *,
    taxonomy: dict[str, list[str]],
    api_key: str,
    model: str,
) -> list[tuple[str, str, str]]:
    payload = {
        "model": model,
        "messages": [
            {
                "role": "system",
                "content": (
                    "Classify low-rating TV reviews into the given taxonomy. "
                    "Use meaning, not keyword matching. Pick one main complaint, especially the one explaining low rating/return. "
                    "primary must be an exact taxonomy key; secondary must be exact under that primary. Do not invent labels. "
                    "If unclear/no complaint/no fit: primary='Other', secondary='Please add comments'. "
                    "Strict boundaries: Cannot turn on only for explicit no power/no response/cannot wake/black and cannot start; "
                    "not for screen damage, no picture with sound, freeze, app/network/remote failure, or generic broken/died/failed. "
                    "OTA_failure only for update process failure/stuck/detect/download/install/complete failure; "
                    "if issue happened after a completed update, classify the actual resulting issue. "
                    "Return concise Chinese reason. Return only JSON array: "
                    "[{\"id\":0,\"primary\":\"...\",\"secondary\":\"...\",\"reason\":\"...\"}]."
                ),
            },
            {
                "role": "user",
                "content": json.dumps(
                    {
                        "taxonomy": taxonomy,
                        "reviews": rows,
                    },
                    ensure_ascii=False,
                ),
            },
        ],
        "temperature": 0,
        "max_tokens": read_int_env("REVIEW_ANALYSIS_CLASSIFY_MAX_TOKENS", default=5000, minimum=200, maximum=12000),
        "max_completion_tokens": read_int_env("REVIEW_ANALYSIS_CLASSIFY_MAX_TOKENS", default=5000, minimum=200, maximum=12000),
        "reasoning_split": True,
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }

    response = post_chat_completion(payload, headers=headers, action="语义分类")
    if response.status_code == 429:
        raise RuntimeError("OpenAI API 额度不足或账单不可用，无法进行语义分类。")
    if response.status_code >= 400:
        raise RuntimeError(format_api_error(response, "语义分类"))
    response.raise_for_status()

    response_data = response.json()
    content = response_message_text(response_data)
    parsed = parse_semantic_items(content, rows)
    by_id = {
        int(item.get("id")): validate_semantic_classification(item, taxonomy)
        for item in parsed
        if isinstance(item, dict) and "id" in item
    }
    return [
        by_id.get(
            row["id"],
            ("Other", "Please add comments", "API 未返回有效分类结果"),
        )
        for row in rows
    ]


def validate_semantic_classification(
    item: dict,
    taxonomy: dict[str, list[str]],
) -> tuple[str, str, str]:
    primary = str(item.get("primary", "")).strip()
    secondary = str(item.get("secondary", "")).strip()
    reason = str(item.get("reason", "")).strip()
    fallback_secondary = (
        "Please add comments"
        if "Please add comments" in taxonomy.get("Other", [])
        else (taxonomy.get("Other", ["Other"])[0] if taxonomy.get("Other") else "Other")
    )
    if primary not in taxonomy:
        return "Other", fallback_secondary, reason or "返回的一级分类不在分类表中"
    if secondary not in taxonomy[primary]:
        return "Other", fallback_secondary, reason or "返回的二级分类不在分类表中"
    return primary, secondary, reason


def truncate_text(value, limit: int) -> str:
    if pd.isna(value):
        return ""
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text if len(text) <= limit else text[: limit - 1] + "…"


def classify_review_keyword(row: pd.Series, issue_rules: list[tuple[str, list[str]]]) -> tuple[str, str]:
    text = normalize_match_text(f"{row.get('标题', '')} {row.get('评论内容', '')}")
    scores = []
    for category, phrases in issue_rules:
        score, matched_phrase = score_category(text, phrases)
        if score:
            scores.append((score, category, matched_phrase))
    if not scores:
        return "Other", "Other"
    scores.sort(key=lambda item: (-item[0], item[1]))
    return scores[0][1], scores[0][2]


def score_category(text: str, phrases: list[str]) -> tuple[int, str]:
    best_score = 0
    best_phrase = ""
    for phrase in phrases:
        normalized = normalize_match_text(phrase)
        if not normalized:
            continue
        if normalized in text:
            score = 8 + len(normalized.split())
            if score > best_score:
                best_score = score
                best_phrase = str(phrase).strip()
            continue

        tokens = meaningful_tokens(normalized)
        if not tokens:
            continue
        matches = sum(1 for token in tokens if token in text)
        if matches >= max(1, min(2, len(tokens))):
            score = matches
            if score > best_score:
                best_score = score
                best_phrase = str(phrase).strip()
    return best_score, best_phrase


def normalize_match_text(value: str) -> str:
    text = str(value).lower()
    text = text.replace("wi-fi", "wifi")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def meaningful_tokens(value: str) -> list[str]:
    return [
        token
        for token in value.split()
        if len(token) >= 3 and token not in STOPWORDS
    ]


def make_excerpt(value, limit: int = 220) -> str:
    if pd.isna(value):
        return ""
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text if len(text) <= limit else text[: limit - 1] + "…"


def make_issue_summary(row: pd.Series) -> str:
    body = make_excerpt(row.get("评论内容", ""))
    if body:
        return body

    title = make_excerpt(row.get("标题", ""))
    if title and not re.fullmatch(r"\d+(?:\.\d+)?\s+out of 5 stars", title, flags=re.I):
        return title

    return "原始评论无正文"


def translate_summaries(values: list[str]) -> list[str]:
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    if not api_key:
        safe_print(
            "WARNING: OPENAI_API_KEY 未配置，问题摘要中文列将留空。可在 .env 中添加 OPENAI_API_KEY。",
            stream=sys.stderr,
        )
        return [""] * len(values)

    model = os.getenv("REVIEW_ANALYSIS_TRANSLATE_MODEL", "gpt-4o-mini").strip()
    batch_size = read_int_env("REVIEW_ANALYSIS_TRANSLATE_BATCH_SIZE", default=10, minimum=1, maximum=50)
    cache = load_translation_cache()
    translations: list[str | None] = [None] * len(values)
    pending: list[tuple[int, str]] = []

    for index, value in enumerate(values):
        key = text_cache_key(value)
        cached = cache.get(key)
        if cached is not None:
            translations[index] = cached
        else:
            pending.append((index, value))

    cached_count = len(values) - len(pending)
    if cached_count:
        safe_print(f"摘要翻译缓存: 已复用 {cached_count}/{len(values)} 条")

    completed = cached_count
    for batch_items in chunked(pending, batch_size):
        batch_values = [value for _, value in batch_items]
        batch_results = translate_batch(batch_values, api_key=api_key, model=model)
        for (index, value), translated in zip(batch_items, batch_results):
            translations[index] = translated
            cache[text_cache_key(value)] = translated
        save_translation_cache(cache)
        completed += len(batch_items)
        safe_print(f"摘要翻译进度: {completed}/{len(values)} 条")

    return [item or "" for item in translations]


def translate_batch(values: list[str], *, api_key: str, model: str) -> list[str]:
    if not values:
        return []

    payload = {
        "model": model,
        "messages": [
            {
                "role": "system",
                "content": (
                    "Translate product review issue summaries into concise Simplified Chinese. "
                    "Preserve product terms, app names, model names, and technical abbreviations. "
                    "Return only a JSON array of strings, with the same length and order as input."
                ),
            },
            {
                "role": "user",
                "content": json.dumps(values, ensure_ascii=False),
            },
        ],
        "temperature": 0,
        "max_tokens": read_int_env("REVIEW_ANALYSIS_TRANSLATE_MAX_TOKENS", default=3000, minimum=100, maximum=8000),
        "max_completion_tokens": read_int_env("REVIEW_ANALYSIS_TRANSLATE_MAX_TOKENS", default=3000, minimum=100, maximum=8000),
        "reasoning_split": True,
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }

    try:
        response = post_chat_completion(payload, headers=headers, action="问题摘要翻译")
        if response.status_code == 429:
            safe_print(
                "WARNING: OpenAI API 额度不足或账单不可用，问题摘要中文列将留空。",
                stream=sys.stderr,
            )
            return [""] * len(values)
        if response.status_code >= 400:
            safe_print(
                f"WARNING: {format_api_error(response, '问题摘要翻译')}",
                stream=sys.stderr,
            )
            return [""] * len(values)
        response.raise_for_status()
        response_data = response.json()
        content = response_message_text(response_data)
        translated = json.loads(extract_json_array(content))
        if isinstance(translated, list) and len(translated) == len(values):
            return [str(item) for item in translated]
    except Exception as exc:  # noqa: BLE001
        safe_print(f"WARNING: translation failed for one batch: {exc}", stream=sys.stderr)

    return [""] * len(values)


def strip_code_fence(value: str) -> str:
    text = value.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.I)
        text = re.sub(r"\s*```$", "", text)
    return text.strip()


def extract_json_array(value: str) -> str:
    text = strip_code_fence(value)
    text = re.sub(r"<think>.*?</think>", "", text, flags=re.I | re.S).strip()
    if text.startswith("[") and text.endswith("]"):
        return text

    start = text.find("[")
    end = text.rfind("]")
    if start != -1 and end != -1 and end > start:
        return text[start : end + 1]

    preview = re.sub(r"\s+", " ", text)[:500]
    raise ValueError(f"API 返回内容不是 JSON 数组，无法解析。返回开头: {preview or '<empty>'}")


def parse_semantic_items(content: str, rows: list[dict]) -> list[dict]:
    json_text = extract_json_array(content)
    try:
        parsed = json.loads(json_text)
        if isinstance(parsed, list):
            return [item for item in parsed if isinstance(item, dict)]
    except json.JSONDecodeError:
        recovered = recover_semantic_items(json_text)
        if recovered:
            safe_print(
                "WARNING: API 返回的 JSON 格式不标准，已用宽松解析恢复本批结果。",
                stream=sys.stderr,
            )
            return recovered

    preview = re.sub(r"\s+", " ", json_text)[:500]
    safe_print(
        f"WARNING: API 返回结果无法解析，本批 {len(rows)} 条将标记为 Other。返回开头: {preview}",
        stream=sys.stderr,
    )
    return [
        {
            "id": row["id"],
            "primary": "Other",
            "secondary": "Please add comments",
            "reason": "API 返回格式异常，未能稳定解析分类结果",
        }
        for row in rows
    ]


def recover_semantic_items(json_text: str) -> list[dict]:
    recovered: list[dict] = []
    for obj_text in re.findall(r"\{.*?\}", json_text, flags=re.S):
        id_match = re.search(r'"id"\s*:\s*(\d+)', obj_text)
        primary_match = re.search(r'"primary"\s*:\s*"([^"]*)"', obj_text)
        secondary_match = re.search(r'"secondary"\s*:\s*"([^"]*)"', obj_text)
        reason_match = re.search(r'"reason"\s*:\s*"(.*)"\s*$', obj_text.strip()[:-1], flags=re.S)
        if not (id_match and primary_match and secondary_match):
            continue
        recovered.append(
            {
                "id": int(id_match.group(1)),
                "primary": primary_match.group(1).strip(),
                "secondary": secondary_match.group(1).strip(),
                "reason": clean_recovered_json_string(reason_match.group(1)) if reason_match else "API 返回格式已修复",
            }
        )
    return recovered


def clean_recovered_json_string(value: str) -> str:
    return (
        value.replace('\\"', '"')
        .replace("\\n", " ")
        .replace("\\r", " ")
        .replace("\\t", " ")
        .strip()
    )


def response_message_text(response_data: dict) -> str:
    choices = response_data.get("choices") or []
    if not choices:
        raise ValueError(f"API 返回中没有 choices。返回内容: {preview_json(response_data)}")

    choice = choices[0]
    message = choice.get("message") or {}
    content = message.get("content")
    if isinstance(content, str) and content.strip():
        return content.strip()
    if isinstance(content, list):
        text_parts = [
            str(part.get("text", ""))
            for part in content
            if isinstance(part, dict) and part.get("type") in {"text", "output_text"}
        ]
        text = "\n".join(part for part in text_parts if part.strip()).strip()
        if text:
            return text

    diagnostics = {
        "finish_reason": choice.get("finish_reason"),
        "message_keys": sorted(message.keys()),
        "usage": response_data.get("usage"),
    }
    raise ValueError(
        "API 返回正文为空，可能是输出 token 上限太低或模型只返回了 thinking。"
        f"诊断信息: {preview_json(diagnostics)}"
    )


def post_chat_completion(payload: dict, *, headers: dict[str, str], action: str) -> httpx.Response:
    timeout_seconds = read_int_env("REVIEW_ANALYSIS_API_TIMEOUT_SECONDS", default=240, minimum=30, maximum=900)
    retries = read_int_env("REVIEW_ANALYSIS_API_RETRIES", default=3, minimum=1, maximum=8)
    last_error: Exception | None = None

    for attempt in range(1, retries + 1):
        try:
            with httpx.Client(timeout=timeout_seconds) as client:
                return client.post(
                    chat_completions_url(),
                    headers=headers,
                    json=payload,
                )
        except httpx.TimeoutException as exc:
            last_error = exc
            if attempt >= retries:
                break
            wait_seconds = min(10 * attempt, 30)
            safe_print(
                f"WARNING: {action} API 第 {attempt}/{retries} 次请求超时，"
                f"{wait_seconds} 秒后重试...",
                stream=sys.stderr,
            )
            time.sleep(wait_seconds)

    raise RuntimeError(f"{action} API 请求超时，已重试 {retries} 次。最后错误: {last_error}")


def semantic_cache_path() -> Path:
    return OUTPUT_DIR / "review_semantic_cache.json"


def load_semantic_cache() -> dict[str, list[str]]:
    path = semantic_cache_path()
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            return {
                str(key): value
                for key, value in data.items()
                if isinstance(value, list) and len(value) >= 2
            }
    except Exception as exc:  # noqa: BLE001
        safe_print(f"WARNING: 读取语义分类缓存失败，将忽略缓存: {exc}", stream=sys.stderr)
    return {}


def save_semantic_cache(cache: dict[str, list[str]]) -> None:
    path = semantic_cache_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def translation_cache_path() -> Path:
    return OUTPUT_DIR / "review_translation_cache.json"


def load_translation_cache() -> dict[str, str]:
    path = translation_cache_path()
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            return {str(key): str(value) for key, value in data.items()}
    except Exception as exc:  # noqa: BLE001
        safe_print(f"WARNING: 读取摘要翻译缓存失败，将忽略缓存: {exc}", stream=sys.stderr)
    return {}


def save_translation_cache(cache: dict[str, str]) -> None:
    path = translation_cache_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def review_cache_key(row: dict) -> str:
    raw = str(row.get("cache_id") or "")
    if not raw:
        payload = {
            key: row.get(key, "")
            for key in ["model", "channel", "rating", "date", "title", "review", "summary"]
        }
        raw = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def text_cache_key(value: str) -> str:
    normalized = re.sub(r"\s+", " ", str(value)).strip()
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def make_review_stable_id(row: pd.Series) -> str:
    parts = [
        str(row.get("机型ID", "")),
        str(row.get("来源站点", "")),
        str(row.get("来源URL", "")),
        str(row.get("评分", "")),
        str(row.get("评论日期", "")),
        str(row.get("用户", "")),
        normalize_cache_text(row.get("标题", "")),
        normalize_cache_text(row.get("评论内容", "")),
    ]
    return "|".join(parts)


def normalize_cache_text(value) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def preview_json(value) -> str:
    text = json.dumps(value, ensure_ascii=False, default=str)
    return text if len(text) <= 800 else text[:797] + "..."


def chat_completions_url() -> str:
    base_url = (
        os.getenv("OPENAI_BASE_URL")
        or os.getenv("OPENAI_API_BASE_URL")
        or "https://api.openai.com/v1"
    ).strip()
    return f"{base_url.rstrip('/')}/chat/completions"


def format_api_error(response: httpx.Response, action: str) -> str:
    detail = response.text.strip()
    if len(detail) > 800:
        detail = detail[:797] + "..."
    return (
        f"{action} API 请求失败：HTTP {response.status_code} {response.reason_phrase}; "
        f"URL={response.request.url}; 返回内容={detail or '<empty>'}"
    )


def read_int_env(name: str, *, default: int, minimum: int, maximum: int) -> int:
    raw_value = os.getenv(name, "").strip()
    if not raw_value:
        return default
    try:
        value = int(raw_value)
    except ValueError:
        return default
    return max(minimum, min(maximum, value))


def chunked(values: list[str], size: int):
    for start in range(0, len(values), size):
        yield values[start : start + size]


def write_analysis(
    df: pd.DataFrame,
    all_reviews: pd.DataFrame,
    output_path: Path,
    source_path: Path,
    dedupe_stats: dict[str, int],
    issue_rules: list[tuple[str, list[str]]],
    category_path: Path | None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if df.empty:
            empty = pd.DataFrame(
                [["未发现 3 星及以下评论", str(source_path)]],
                columns=["说明", "来源文件"],
            )
            empty.to_excel(writer, sheet_name="无低分评论", index=False)
            return

        total_reviews_by_model = all_reviews.groupby("机型ID", sort=False).size().to_dict()
        all_reviews_by_model = {
            model_id: model_df.copy()
            for model_id, model_df in all_reviews.groupby("机型ID", sort=False)
        }
        for model_id, model_df in df.groupby("机型ID", sort=False):
            write_model_sheet(
                writer,
                str(model_id),
                model_df,
                all_reviews_by_model.get(model_id, pd.DataFrame()),
                source_path,
                issue_rules,
                total_reviews_by_model.get(model_id, len(model_df)),
            )

        format_workbook(writer.book)


def write_model_sheet(
    writer: pd.ExcelWriter,
    model_id: str,
    model_df: pd.DataFrame,
    all_model_reviews: pd.DataFrame,
    source_path: Path,
    issue_rules: list[tuple[str, list[str]]],
    total_reviews_all_ratings: int,
) -> None:
    total = len(model_df)
    counts = Counter(model_df["问题分类"])
    summary = pd.DataFrame(
        [
            {
                "问题分类": category,
                "问题数": count,
                "问题占比": count / total if total else 0,
                "代表关键词": representative_keywords(category, issue_rules),
            }
            for category, count in counts.most_common()
        ]
    )

    detail_cols = [
        "机型名称",
        "Channel",
        "评分",
        "评论日期",
        "问题分类",
        "分类理由",
        "评论内容",
        "问题摘要中文",
    ]
    detail_cols = [col for col in detail_cols if col in model_df.columns]
    details = model_df.copy()
    details["_sort_date"] = details["评论日期"].map(parse_review_date)
    details["评论日期"] = details["评论日期"].map(format_review_date)
    details["问题分类"] = details["二级问题分类"]
    details = details.sort_values(
        by="_sort_date",
        ascending=False,
        na_position="last",
        kind="stable",
    )[detail_cols]

    sheet_name = safe_sheet_name(model_id)
    channel_metrics = build_channel_metrics_table(all_model_reviews)
    header = pd.DataFrame(
        [
            [f"{model_id} 低分评论问题分析"],
            [f"来源文件: {source_path.name}"],
            [f"总评论数（1-5星）: {total_reviews_all_ratings}"],
            [f"3星及以下评论数: {total}"],
        ]
    )
    metrics_title = pd.DataFrame([["渠道评分与评论数"]])
    summary_title = pd.DataFrame([["低分问题分类汇总"]])
    header.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=0)
    metrics_title.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=5)
    channel_metrics.to_excel(writer, sheet_name=sheet_name, index=False, startrow=6)
    summary_start = 9 + len(channel_metrics)
    summary_title.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=summary_start)
    summary.to_excel(writer, sheet_name=sheet_name, index=False, startrow=summary_start + 1)
    details.to_excel(writer, sheet_name=sheet_name, index=False, startrow=summary_start + 4 + len(summary))


def build_channel_metrics_table(all_model_reviews: pd.DataFrame) -> pd.DataFrame:
    columns = ["渠道", "尺寸", "机型名称", "总评分", "评论数"]
    if all_model_reviews.empty:
        return pd.DataFrame(columns=columns)

    reviews = all_model_reviews.copy()
    reviews["_rating_num"] = reviews["评分"].map(parse_rating) if "评分" in reviews.columns else None
    reviews["_overall_rating_num"] = reviews["整体评分"].map(parse_rating) if "整体评分" in reviews.columns else None
    reviews["_size_display"] = reviews.apply(extract_size, axis=1)
    channel = reviews["Channel"].fillna("").astype(str).str.upper() if "Channel" in reviews.columns else pd.Series("", index=reviews.index)
    rows = []

    amz = reviews[channel.eq("AMZ")]
    if not amz.empty:
        rows.append(
            {
                "渠道": "AMZ",
                "尺寸": "全部",
                "机型名称": display_model_name(amz),
                "总评分": overall_or_average_rating(amz),
                "评论数": len(amz),
            }
        )

    bby = reviews[channel.eq("BBY")]
    if not bby.empty:
        for size, size_df in bby.groupby("_size_display", sort=False):
            rows.append(
                {
                    "渠道": "BBY",
                    "尺寸": size or "未知",
                    "机型名称": display_model_name(size_df),
                    "总评分": overall_or_average_rating(size_df),
                    "评论数": len(size_df),
                }
            )

    return pd.DataFrame(rows, columns=columns)


def display_model_name(df: pd.DataFrame) -> str:
    for column in ("机型名称", "机型ID"):
        if column not in df.columns:
            continue
        values = df[column].dropna().astype(str).str.strip()
        values = values[values.ne("")]
        if not values.empty:
            return values.iloc[0]
    return ""


def overall_or_average_rating(df: pd.DataFrame) -> float | None:
    overall_values = df["_overall_rating_num"].dropna() if "_overall_rating_num" in df.columns else pd.Series(dtype=float)
    if not overall_values.empty:
        return round(float(overall_values.iloc[0]), 1)
    rating_values = df["_rating_num"].dropna() if "_rating_num" in df.columns else pd.Series(dtype=float)
    if not rating_values.empty:
        return round(float(rating_values.mean()), 1)
    return None


def representative_keywords(category: str, issue_rules: list[tuple[str, list[str]]]) -> str:
    for name, keywords in issue_rules:
        if name == category:
            return ", ".join(keywords[:8])
    return ""


def safe_sheet_name(name: str, max_len: int = 31) -> str:
    invalid = set(r"[]:*?/\\")
    cleaned = "".join(c if c not in invalid else "_" for c in name).strip() or "sheet"
    return cleaned[:max_len]


def format_workbook(workbook) -> None:
    hisense_green = "00979B"
    header_fill = PatternFill("solid", fgColor=hisense_green)
    thin_side = Side(style="thin", color="B7C9CC")
    table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    body_font = Font(name="Microsoft YaHei", size=11)
    header_font = Font(name="Microsoft YaHei", size=11, color="FFFFFF", bold=True)
    title_font = Font(name="Microsoft YaHei", size=14, bold=True, color=hisense_green)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A7"
        if sheet.max_column > 1:
            sheet.merge_cells(
                start_row=1,
                start_column=1,
                end_row=1,
                end_column=sheet.max_column,
            )
            sheet.cell(row=1, column=1).alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        for row in sheet.iter_rows():
            for cell in row:
                cell.font = body_font
                set_cell_alignment(cell, vertical="center", wrap_text=True)
                if cell.row == 1 and cell.column == 1:
                    cell.font = title_font
                if is_analysis_header_cell(cell.value):
                    cell.fill = header_fill
                    cell.font = header_font

        sheet.cell(row=1, column=1).alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        for column in range(1, sheet.max_column + 1):
            letter = get_column_letter(column)
            if column == 1:
                sheet.column_dimensions[letter].width = 14
            else:
                sheet.column_dimensions[letter].width = column_width(sheet, column)

        for row in range(1, sheet.max_row + 1):
            sheet.row_dimensions[row].height = 24

        center_columns_by_header(sheet, {"渠道", "尺寸", "机型名称", "总评分", "评论数", "问题分类", "问题数", "问题占比"})
        apply_column_number_formats(sheet)
        apply_analysis_table_borders(sheet, table_border)


def center_columns_by_header(sheet, headers_to_center: set[str]) -> None:
    for row in sheet.iter_rows():
        for header_cell in row:
            if header_cell.value not in headers_to_center:
                continue
            header_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            for cell in sheet.iter_cols(
                min_col=header_cell.column,
                max_col=header_cell.column,
                min_row=header_cell.row + 1,
                max_row=sheet.max_row,
            ):
                for data_cell in cell:
                    set_cell_alignment(
                        data_cell,
                        horizontal="center",
                        vertical="center",
                        wrap_text=True,
                    )


def set_cell_alignment(cell, *, horizontal=None, vertical=None, wrap_text=None) -> None:
    alignment = copy(cell.alignment)
    if horizontal is not None:
        alignment.horizontal = horizontal
    if vertical is not None:
        alignment.vertical = vertical
    if wrap_text is not None:
        alignment.wrap_text = wrap_text
    cell.alignment = alignment


def apply_column_number_formats(sheet) -> None:
    for row in sheet.iter_rows():
        headers = {cell.column: cell.value for cell in row if isinstance(cell.value, str)}
        if not headers:
            continue

        data_start = row[0].row + 1
        data_end = table_data_end_row(sheet, data_start)
        for column, header in headers.items():
            if "百分比" in header or "占比" in header or "比例" in header:
                for data_cell in sheet.iter_cols(
                    min_col=column,
                    max_col=column,
                    min_row=data_start,
                    max_row=data_end,
                ):
                    for cell in data_cell:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "0.0%"
            elif header in {"问题数", "评论数", "低分评论总数"}:
                for data_cell in sheet.iter_cols(
                    min_col=column,
                    max_col=column,
                    min_row=data_start,
                    max_row=data_end,
                ):
                    for cell in data_cell:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "0"
            elif header in {"评分", "总评分"}:
                for data_cell in sheet.iter_cols(
                    min_col=column,
                    max_col=column,
                    min_row=data_start,
                    max_row=data_end,
                ):
                    for cell in data_cell:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "0.0"


def table_data_end_row(sheet, data_start: int) -> int:
    row_idx = data_start
    while row_idx <= sheet.max_row:
        values = [sheet.cell(row=row_idx, column=col).value for col in range(1, sheet.max_column + 1)]
        if all(value is None for value in values):
            return row_idx - 1
        row_idx += 1
    return sheet.max_row


def apply_used_range_borders(sheet, border: Border) -> None:
    for row in sheet.iter_rows(
        min_row=1,
        max_row=sheet.max_row,
        min_col=1,
        max_col=sheet.max_column,
    ):
        for cell in row:
            cell.border = border


def is_analysis_header_cell(value) -> bool:
    return value in {
        "问题分类",
        "问题数",
        "问题占比",
        "代表关键词",
        "渠道",
        "尺寸",
        "总评分",
        "评论数",
        "机型ID",
        "机型名称",
        "Channel",
        "评分",
        "评论日期",
        "分类理由",
        "评论内容",
        "问题摘要中文",
        "低分评论总数",
    }


def apply_analysis_table_borders(sheet, border: Border) -> None:
    for row in sheet.iter_rows():
        if not any(is_analysis_header_cell(cell.value) for cell in row):
            continue
        populated_columns = [cell.column for cell in row if cell.value is not None]
        if not populated_columns:
            continue
        start_column = min(populated_columns)
        end_column = max(populated_columns)
        end_row = table_data_end_row(sheet, row[0].row + 1)
        for cells in sheet.iter_rows(
            min_row=row[0].row,
            max_row=end_row,
            min_col=start_column,
            max_col=end_column,
        ):
            for cell in cells:
                cell.border = border


def column_width(sheet, column: int) -> int:
    values = [sheet.cell(row=row, column=column).value for row in range(1, min(sheet.max_row, 80) + 1)]
    max_len = max((len(str(value)) for value in values if value is not None), default=8)
    return max(10, min(max_len + 2, 45))


def safe_print(value: str, *, stream=None) -> None:
    target = stream or sys.stdout
    try:
        print(value, file=target)
    except UnicodeEncodeError:
        encoded = value.encode(target.encoding or "utf-8", errors="backslashreplace")
        print(encoded.decode(target.encoding or "utf-8"), file=target)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:  # noqa: BLE001
        safe_print(f"ERROR: {exc}", stream=sys.stderr)
        raise
